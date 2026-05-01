from PyQt6 import uic
import os
from openpyxl import load_workbook
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QMessageBox, QFileDialog, QTableWidget, QComboBox, QHeaderView, QTableWidgetItem
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QFont
from db.conexion import ConexionMysql
from db.querys import Query
from datetime import datetime


class Inventory(QMainWindow):
    def __init__(self):
        super().__init__()
        self.main = uic.loadUi('views/inventario.ui')
        self.main.show()
        self.db = ConexionMysql()
        self.error = QMessageBox(self)
        self.main.fechaRegistro.setDate(datetime.now().date())
        self.main.fechaUp.setDate(datetime.now().date())
        self.main.botonRegistrar.clicked.connect(self.registrarProducto)
        self.main.botonCargar.clicked.connect(self.abrirExcel)
        self.main.botonRListado.clicked.connect(self.registrarListado)
        self.main.botonLimpiar.clicked.connect(self.limpiarTodo)
        self.main.listadoProductos.setEditable(True)
        self.main.listadoProductos.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.main.listadoProductosActualizacion.setEditable(True)
        self.main.listadoProductosActualizacion.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.timer = QTimer()
        self.timer.setInterval(800)
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(self.ejecutarBusqueda)
        self.main.listadoProductos.lineEdit().textChanged.connect(self.realizarBusqueda)
        self.main.listadoProductosActualizacion.lineEdit().textChanged.connect(self.realizarBusqueda)
        self.main.listadoProductosActualizacion.activated.connect(self.actualizarProducto)
        self.main.botonLimpiarV.clicked.connect(self.limpiarTodo)
        self.main.botonAgregar.clicked.connect(self.agregarVenta)
        self.main.botonEliminar.clicked.connect(self.borrarRegistro)
        self.main.botonActualizar.clicked.connect(self.actualizar)
        self.main.botonLimpiarA.clicked.connect(self.limpiar)
        self.main.botonVender.clicked.connect(self.vender)
        self.showTVenta()
        self.showTProducto()
        self.seleccionando = False
        self.total = 0
        
        
    def showTVenta(self):
        columns = ['ID','CANTIDAD','DESC. DEL PRODUCTO', 'MEDIDA', 'PRECIO VENTA', 'SUBTOTAL']
        self.main.tablaVenta.setFont(QFont("Arial", 12))
        self.main.tablaVenta.setColumnCount(len(columns))
        for column, name in enumerate(columns):
            self.main.tablaVenta.setHorizontalHeaderItem(column, QTableWidgetItem(name))
        self.main.tablaVenta.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header_style = """
        QHeaderView::section {
            font-family: "Arial";
            font-size: 12pt;
            font-weight: bold;
            background-color: rgb(255, 255, 255);
        }
        QTableWidget{
            background-color: rgb(255, 255, 255);
        }
        """
        self.main.tablaVenta.setStyleSheet(header_style)
        
    def showTProducto(self):
        columns = ['CANTIDAD', 'DESC. DEL PRODUCTO', 'MEDIDA', 'PRECIO COSTO', 'PRECIO VENTA', 'FECHA']
        self.main.tablaProducto.setFont(QFont("Arial", 12))
        self.main.tablaProducto.setColumnCount(len(columns))
        for column, name in enumerate(columns):
            self.main.tablaProducto.setHorizontalHeaderItem(column, QTableWidgetItem(name))
        self.main.tablaProducto.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.main.tablaProducto.setShowGrid(True)
        self.main.tablaProducto.setGridStyle(Qt.PenStyle.SolidLine)
        header_style = """
        QHeaderView::section {
            font-family: "Arial";
            font-size: 12pt;
            font-weight: bold;
            background-color: rgb(255, 255, 255);
        }
        QTableWidget{
            background-color: rgb(255, 255, 255);
        }
        
        
        """
        self.main.tablaProducto.setStyleSheet(header_style)
        
    def registrarProducto(self):
        query = Query()
        try:
            cantidad = int(self.main.textCantidad.text())
            descripcion = str(self.main.textDescripcion.text())
            medida = str(self.main.textMedicion.text())
            precioCosto = float(self.main.textCosto.text())
            precioVenta = float(self.main.textVenta.text())
            fecha = self.main.fechaRegistro.date().toString("yyyy-MM-dd")
            fila = self.main.tablaProducto.rowCount()
            query.insertarProducto(cantidad, descripcion, medida, precioCosto, precioVenta, fecha)
            self.limpiar()
            self.db.close_connection()
            self.main.tablaProducto.insertRow(fila)
            self.main.tablaProducto.setItem(fila, 0, QTableWidgetItem(str(cantidad)))
            self.main.tablaProducto.setItem(fila, 1, QTableWidgetItem(descripcion))
            self.main.tablaProducto.setItem(fila, 2, QTableWidgetItem(medida))
            self.main.tablaProducto.setItem(fila, 3, QTableWidgetItem(str(precioCosto)))
            self.main.tablaProducto.setItem(fila, 4, QTableWidgetItem(str(precioVenta)))
            self.main.tablaProducto.setItem(fila, 5, QTableWidgetItem(fecha))            
        except Exception as e:
            self.error.critical(self, 'Error', f"ERROR: {e}")
            
    def abrirExcel(self):
        folder = QFileDialog()
        try: 
            folder_path, __= folder.getOpenFileName(None, 'ABRIR ARCHIVO', '', 'xlsx (*.xlsx)')
            self.cargarExcel(folder_path)
        except Exception as e:
            self.error.critical(self, 'Error', f"ERROR: {e}")    
    
    def cargarExcel(self, ruta):
        workbook = load_workbook(filename=ruta)
        hoja = workbook.active
        fecha = self.main.fechaRegistro.date().toString("yyyy-MM-dd")
        try: 
            for fila in hoja.iter_rows(values_only=True):
                fila_index = self.main.tablaProducto.rowCount()
                self.main.tablaProducto.insertRow(fila_index)
                self.main.tablaProducto.setItem(fila_index, 0, QTableWidgetItem(str(fila[0])))
                self.main.tablaProducto.setItem(fila_index, 1, QTableWidgetItem(str(fila[1].upper())))
                self.main.tablaProducto.setItem(fila_index, 2, QTableWidgetItem(str(fila[2].upper())))
                self.main.tablaProducto.setItem(fila_index, 3, QTableWidgetItem(str(fila[3])))
                self.main.tablaProducto.setItem(fila_index, 4, QTableWidgetItem(str(fila[3])))
                self.main.tablaProducto.setItem(fila_index, 5, QTableWidgetItem(fecha))
        except Exception as e:
            self.error.critical(self, 'Error', f"ERROR: {e}")  
            
    def registrarListado(self):
        query = Query()
        message_box = QMessageBox(self)
        message_box.setWindowTitle("Confirmación")
        message_box.setText("¿Estás seguro de que deseas continuar?")
        message_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        message_box.setIcon(QMessageBox.Icon.Question)
        response = message_box.exec()
        try:
            if response == QMessageBox.StandardButton.Yes:
                for fila in range(self.main.tablaProducto.rowCount()):
                    cantidad = self.main.tablaProducto.item(fila,0)
                    descripcion = self.main.tablaProducto.item(fila,1)
                    medida = self.main.tablaProducto.item(fila,2)
                    precioCosto = self.main.tablaProducto.item(fila,3)
                    precioVenta = self.main.tablaProducto.item(fila,4)
                    fecha = self.main.tablaProducto.item(fila,5)
                    query.insertarProducto(int(float(cantidad.text())), descripcion.text(), medida.text(), float(precioCosto.text()), float(precioVenta.text()), fecha.text())
                self.db.close_connection()
        except Exception as e:
            self.error.critical(self, 'Error', f"ERROR: {e}")  
    
    
    def realizarBusqueda(self):
        self.timer.start()
    
    def ejecutarBusqueda(self):
        texto = self.main.listadoProductos.currentText()
        texto2 = self.main.listadoProductosActualizacion.currentText()
        if texto.strip():
            self.buscarProductos(texto)
        elif texto2.strip():
            self.buscarProductoActualizacion(texto2)
            
        """self.buscarProductos(texto)
        self.timer.stop()"""
    
    def buscarProductoActualizacion(self, text):
        query = Query()
        if text == "":
            return
        try:
            protuctos = query.seleccionarProducto(text)
            self.main.listadoProductosActualizacion.blockSignals(True)
            self.main.listadoProductosActualizacion.clear()
            for informacion in protuctos:
                texto_i = f"{informacion[0]} {informacion[1]} {informacion[2]} {informacion[3]} {informacion[4]} {informacion[5]}"
                self.main.listadoProductosActualizacion.addItem(texto_i, informacion)
            self.main.listadoProductosActualizacion.blockSignals(False)
            if protuctos:    
                self.main.listadoProductosActualizacion.showPopup()
        except Exception as e:
            self.error.critical(self, 'Error', f"ERROR: {e}") 
            
    def buscarProductos(self, texto):
        query = Query()
        if texto == "":
            return

        try:
            listProductos = query.seleccionarProducto(texto)
            self.main.listadoProductos.blockSignals(True)
            #self.main.listadoProductos.clear()
            for informacion in listProductos:
                texto_i = f"{informacion[0]} {informacion[1]} {informacion[2]} {informacion[3]} {informacion[5]}"
                self.main.listadoProductos.addItem(texto_i, informacion)
            self.main.listadoProductos.blockSignals(False)
            if listProductos:
               self.main.listadoProductos.showPopup()
        except Exception as e:
            self.error.critical(self, 'Error', f"ERROR: {e}") 
            
    def actualizarProducto(self, index):
        data = self.main.listadoProductosActualizacion.itemData(index)
        if data:
            try: 
                idp = data[0]
                cantidad = data[1]
                descripcion = data[2]
                medida = data[3]
                precioV= data[4]
                precioC = data[5]
                self.main.idProducto.setText(str(idp))
                self.main.upCantidad.setText(str(cantidad))
                self.main.upDescripcion.setText(descripcion)
                self.main.upMedicion.setText(medida)
                self.main.upCosto.setText(str(precioC))
                self.main.upVenta.setText(str(precioV))
            except Exception as e:
                self.error.critical(self, 'Error', f"ERROR: {e}")
    def actualizar (self):
        query = Query()
        message_box = QMessageBox(self)
        messageInfo = QMessageBox()
        message_box.setWindowTitle("Confirmación")
        message_box.setText("¿Estás seguro de que desea actualizar el producto?")
        message_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        message_box.setIcon(QMessageBox.Icon.Question)
        response = message_box.exec()
        try:
            if response == QMessageBox.StandardButton.Yes:
                idp = int(self.main.idProducto.text())
                cantidad = int( self.main.upCantidad.text())
                descripcion = self.main.upDescripcion.text()
                medida = self.main.upMedicion.text()
                precioV= float(self.main.upVenta.text())
                precioC = float(self.main.upCosto.text())
                fecha = self.main.fechaUp.date().toString("yyyy-MM-dd")
                result = query.actualizarProductos(cantidad,descripcion,medida, precioC, precioV, fecha,idp)
                messageInfo.information(None,"ACTUALIZACION", "Operacion realizada con exito")
                self.main.idProducto.setText("")
                self.main.upCantidad.setText("")
                self.main.upDescripcion.setText("")
                self.main.upMedicion.setText("")
                self.main.upCosto.setText("")
                self.main.upVenta.setText("")
                    
        except Exception as e:
            self.error.critical(self, 'Error', f"ERROR: {e}")
            
    def agregarVenta(self):
        self.timer.stop()
        self.main.listadoProductos.blockSignals(True)
        data = self.main.listadoProductos.currentData()

        #data = self.main.listadoProductos.itemData(index)
        print(data)
        if data:
            try:
                cantidad = self.main.textCantidadV.text()
                idProducto = data[0]
                print(idProducto)
                descripcion = data[2]
                medida = data[3]
                precioVenta = data[5] 
                subTotal = int(cantidad)*precioVenta
                fila = self.main.tablaVenta.rowCount()
                self.main.tablaVenta.insertRow(fila)
                self.main.tablaVenta.setItem(fila, 0, QTableWidgetItem(str(idProducto)))
                self.main.tablaVenta.setItem(fila, 1, QTableWidgetItem(cantidad))
                self.main.tablaVenta.setItem(fila, 2, QTableWidgetItem(descripcion))
                self.main.tablaVenta.setItem(fila, 3, QTableWidgetItem(medida))
                self.main.tablaVenta.setItem(fila, 4, QTableWidgetItem(str(precioVenta)))
                self.main.tablaVenta.setItem(fila, 5, QTableWidgetItem(str(subTotal)))
                self.main.listadoProductos.clear()
                self.main.listadoProductos.clearEditText()
                self.main.listadoProductos.blockSignals(False)
                self.total = self.total +subTotal
                self.main.totalVenta.setText(str(self.total))
            except Exception as e:
                self.error.critical(self, 'Error', f"ERROR: {e}")
            finally: 
                self.main.listadoProductos.blockSignals(False)
        else:
            self.error.critical(self, 'Error', f"ERROR: campo vacio o hubo un error interno")
       
    def vender(self):
        tabla = self.main.tablaVenta
        detalle_producto = []
        for fila in range(tabla.rowCount()):
            try:
                item_id = tabla.item(fila, 0)
                item_cantidad = tabla.item(fila, 1)
                item_subTotal = tabla.item(fila, 5)
                registro = {
                    "idProducto": int(item_id.text()),
                    "cantidad": int(item_cantidad.text()),
                    "sub_total": float(item_subTotal.text())
                }
                detalle_producto.append(registro)
                print(detalle_producto)
            except Exception as e:
                self.error.critical(self, 'Error', f"ERROR: {e}")
                
    def borrarRegistro(self):
        fila = self.main.tablaVenta.currentRow()
        columna = self.main.tablaVenta.item(fila, 4)
        if fila== -1:
            self.error.critical(self, 'Error', f"ERROR: no se selecciono ningun registro o hubo un error interno")  
            return
        try:
            self.total = self.total-float(columna.text())
            self.main.totalVenta.setText(str(self.total))
            self.main.tablaVenta.removeRow(fila)
        except Exception as e: 
            self.error.critical(self, 'Error', f"ERROR: {e}")       

    def limpiar(self):
        self.main.textCantidad.setText("")
        self.main.textDescripcion.setText("")
        self.main.textMedicion.setText("")
        self.main.textCosto.setText("")
        self.main.textVenta.setText("")
        self.main.tablaVenta.setRowCount(0)
        self.main.totalVenta.setText("0.00")
        self.main.idProducto.setText("")
        self.main.upCantidad.setText("")
        self.main.upDescripcion.setText("")
        self.main.upMedicion.setText("")
        self.main.upCosto.setText("")
        self.main.upVenta.setText("")
    
      
        
    def limpiarTodo(self):
        self.timer.stop()
        self.main.listadoProductos.blockSignals(True)
        self.main.listadoProductos.clear()
        self.main.listadoProductos.clearEditText()
        self.main.listadoProductos.blockSignals(False)
        self.main.listadoProductosActualizacion.blockSignals(True)
        self.main.listadoProductosActualizacion.clear()
        self.main.listadoProductosActualizacion.clearEditText()
        self.main.listadoProductosActualizacion.blockSignals(False)
        self.limpiar()
        self.main.tablaProducto.setRowCount(0)